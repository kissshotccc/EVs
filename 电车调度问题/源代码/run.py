import re
import shutil
import numpy as np
import pandas as pd
import torch
from env import EVs_Env
from DQN import DQN
import csv
import openpyxl
from openpyxl import Workbook
from data import get_data, get_model_name
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'


def get_max_model(dir_path):
    pattern = re.compile(r"model_(\d+)\.pth")

    # Initialize the maximum model number
    max_num = -1

    # Traverse through all subdirectories of the parent directory
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            match = pattern.match(file)

            if match:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num

    return max_num

def get_best_model(env,num):
    dir_path = f'./model/best_power_path{num+1}'
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            model_path = os.path.join(root,file)
            model_power = DQN_test(env,num,model_path)
            if model_power == env.max_power:
                save_best_model(model_path,num)
                print("car{}保存成功！".format(num+1))
                return -1
                break
        if model_power != env.max_power:
            return num+1

def save_best_model(model_path,num):
    destination_folder = './model/best_power_path'  #保存模型的位置
    #检查目标文件夹是否存在，如果不存在则创建一个
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)


    file_extension = os.path.splitext(model_path)[1]

    #构造新的文件名
    new_model_path = f'car_{num+1}{file_extension}'
    destination_path = os.path.join(destination_folder,new_model_path)

    #复制文件到目标路径
    if not os.path.isfile(new_model_path):
        shutil.copy(model_path,destination_path)

def save_results(results_detail, results_all):

    # 打开或创建 results.xlsx 文件
    file_name = "../输出结果/results.xlsx"
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 如果表不存在，创建 "调度详情"
    if "调度详情" not in wb.sheetnames:
        ws_detail = wb.create_sheet(title="调度详情")
    else:
        ws_detail = wb["调度详情"]

    # 如果表不存在，创建 "汇总"
    if "汇总" not in wb.sheetnames:
        ws_all = wb.create_sheet(title="汇总")
    else:
        ws_all = wb["汇总"]

    # 清空表
    ws_detail.delete_rows(1, ws_detail.max_row)
    ws_all.delete_rows(1, ws_all.max_row)

    # 将数据写入 Excel
    for data_set in results_detail:
        for row_data in data_set:
            row = []
            for item in row_data:
                if isinstance(item, list) or isinstance(item, tuple):
                    # 将元组或列表作为字符串存入单元格
                    row.append(str(item))
                else:
                    row.append(item)

            # 将行添加到 Excel
            ws_detail.append(row)

    for data_set in results_all:
        for row_data in data_set:
            ws_all.append(row_data)  # 直接追加每一行的子列表

    # 计算第二列的总和
    second_column_sum = 0
    for row in ws_all.iter_rows(min_row=1, max_row=ws_all.max_row, min_col=2, max_col=2):
        for cell in row:
            second_column_sum += cell.value

    # 在最后一行写入 'total' 和第二列的总和
    ws_all.append(['total', second_column_sum])

    # 保存 Excel 文件
    wb.save(file_name)


def DQN_test(env, num, model_path, max_steps=500):
    # 创建DQN对象
    agent = DQN(env)

    agent.model.load_state_dict(torch.load(model_path))
    agent.model.eval()  # 设置模型为评估模式

    # [电动汽车序号、起点序号、终点序号、初始电量、电池容量、截止时间]
    info_list = [num, env.start, env.end, env.init_power, env.max_power, env.dead_line]

    # 电车调度路径 [(s1,s2),(s2,s3)...]
    path_list = []

    # [[s1-s2的剩余电量, s1->s2的剩余截止时间]...]
    power_time_list = []

    # 汇总表 [序号, 到终点时剩余电量, 到终点时剩余时间]
    summary = []
    last_power = 0

    state, _ = env.reset()
    # 循环执行每一步操作
    for step in range(max_steps):
        pre = env.current
        # 使用模型选择动作（不使用epsilon-greedy策略，直接用模型的输出）
        action = agent.take_action(state, 1)
        if action == pre:
            action = agent.take_action(state, 2)

        next_state, reward, terminated, info = env.step(action)

        path_list.append((pre,env.current))
        power_time_list.append([env.current_power,env.remain_time])

        state = next_state
        if terminated:
            if env.current == env.end:
                summary.append([num, env.current_power, env.remain_time])
                last_power = env.current_power
            else:
                summary.append([num, 0, env.remain_time])
                last_power = 0
            break

    global result_detail
    global result_all
    result_detail.append([info_list, path_list, power_time_list])
    result_all.append(summary)

    return last_power


if __name__ == "__main__":
    EVs, env_info = get_data()
    # 生成50轮结果，即50轮总电量，最后取平均值
    res = 0
    result_detail = []
    result_all = []
    for i in range(len(EVs)):
        env = EVs_Env(EVs[i], env_info)
        env.reset()
        # # 加载保存的模型权重
        model_path = f'./model/best_power_path/car_{i+1}.pth'
        DQN_test(env, i,model_path)

    save_results(result_detail, result_all)
