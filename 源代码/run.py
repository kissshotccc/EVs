import csv
import numpy as np
import openpyxl
from openpyxl import Workbook
import pandas as pd
import torch
from env import EVs_Env
from DQN import DQN

import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'


def init_data():
    # 50个节点
    data_50_temp = pd.read_csv('数据集/data_50.csv', header=None)
    data = []
    for i in range(0, 50):
        data.append((data_50_temp.iloc[i][0], data_50_temp.iloc[i][1]))

    # 节点间距离
    distance_50_temp = pd.read_csv('数据集/distance_50.csv', header=None)
    distance = distance_50_temp.values

    # 两节点之间是否为充电路段
    roads_50_temp = pd.read_csv('数据集/roads_50.csv', header=None)
    roads = roads_50_temp.values

    # 两个节点之间的行驶速度
    speed_50_temp = pd.read_csv('数据集/speed_50.csv', header=None)
    speed = speed_50_temp.values

    # 每辆电车的信息
    EVs_50_temp = pd.read_csv('数据集/EVs_50.csv', header=None)
    EVs = {}
    for i in range(0, 50):
        EVs[i] = {}
        EVs[i]['start'] = data.index((EVs_50_temp.iloc[i][0], EVs_50_temp.iloc[i][1]))
        EVs[i]['end'] = data.index((EVs_50_temp.iloc[i][2], EVs_50_temp.iloc[i][3]))
        EVs[i]['init_power'] = EVs_50_temp.iloc[i][4]
        EVs[i]['max_power'] = EVs_50_temp.iloc[i][5]
        EVs[i]['dead_line'] = EVs_50_temp.iloc[i][6]
        EVs[i]['consumption'] = EVs_50_temp.iloc[i][7]

    node_road = np.ones((50, 50))
    # 将主对角线元素设为0
    np.fill_diagonal(node_road, 0)

    env_info = {
        'data': data,
        'distance': distance,
        'speed': speed,
        'roads': roads,
        'node_road': node_road
    }
    return EVs, env_info


def DQN_test(env, num, max_steps=500):
    # 创建DQN对象
    agent = DQN(env)

    # 加载保存的模型权重
    model_path = f'模型/dqn_model_car_{num}.pth'
    agent.model.load_state_dict(torch.load(model_path, weights_only=True))
    agent.model.eval()  # 设置模型为评估模式

    # [电动汽车序号、起点序号、终点序号、初始电量、电池容量、截止时间]
    info_list = [num, env.start, env.end, env.init_power, env.max_power, env.dead_line]

    # 电车调度路径 [(s1,s2),(s2,s3)...]
    path_list = []

    # [[s1-s2的剩余电量, s1->s2的剩余截止时间]...]
    power_time_list = []

    # 汇总表 [序号, 到终点时剩余电量, 到终点时剩余时间]
    summary = []

    state, _ = env.reset()

    # 循环执行每一步操作
    for step in range(max_steps):
        pre = env.current
        # 使用模型选择动作（不使用epsilon-greedy策略，直接用模型的输出）
        action = agent.choose_action(state, epsilon=-1)
        next_state, reward, terminated, info = env.step(action)

        path_list.append((pre,env.current))
        power_time_list.append([env.current_power,env.remain_time])

        state = next_state

        if terminated:
            if env.current == env.end:
                summary.append([num, env.current_power, env.remain_time])
            else:
                summary.append([num, 0, env.remain_time])
            break

    # 保存到csv
    with open('../输出结果/results_detail.csv', 'a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        # 写入数据（追加一行）
        writer.writerow(info_list)
        writer.writerow(path_list)
        writer.writerow(power_time_list)

    with open('../输出结果/results_all.csv', 'a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        # 写入数据（追加一行）
        for row in summary:
            writer.writerow(row)


def save_results(results_detail, results_all):

    # 打开或创建 results.xlsx 文件
    file_name = "../输出结果/results.xlsx"
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 如果表不存在，创建 "调度详情"
    if "调度详情" not in wb.sheetnames:
        ws_detail = wb.create_sheet(title="调度详情")
    else:
        ws_detail = wb["调度详情"]

    # 如果表不存在，创建 "汇总"
    if "汇总" not in wb.sheetnames:
        ws_all = wb.create_sheet(title="汇总")
    else:
        ws_all = wb["汇总"]

    # 清空表
    ws_detail.delete_rows(1, ws_detail.max_row)
    ws_all.delete_rows(1, ws_all.max_row)

    # 将数据写入 Excel
    for data_set in results_detail:
        for row_data in data_set:
            row = []
            for item in row_data:
                if isinstance(item, list) or isinstance(item, tuple):
                    # 将元组或列表作为字符串存入单元格
                    row.append(str(item))
                else:
                    row.append(item)

            # 将行添加到 Excel
            ws_detail.append(row)

    for data_set in results_all:
        for row_data in data_set:
            ws_all.append(row_data)  # 直接追加每一行的子列表

    # 计算第二列的总和
    second_column_sum = 0
    for row in ws_all.iter_rows(min_row=1, max_row=ws_all.max_row, min_col=2, max_col=2):
        for cell in row:
            second_column_sum += cell.value

    # 在最后一行写入 'total' 和第二列的总和
    ws_all.append(['total', second_column_sum])

    # 保存 Excel 文件
    wb.save(file_name)


if __name__ == '__main__':
    EVs, env_info = init_data()
    # print(json.dumps(EVs_50, indent=2))


    # 生成50轮结果，即50轮总电量，最后取平均值
    res = np.array([])
    for k in range(50):

        # 测试50辆车
        for i in range(len(EVs)):
            env = EVs_Env(EVs[i], env_info)
            env.reset()
            DQN_test(env, i)

        # 读取 CSV 文件 
        df = pd.read_csv('../输出结果/results_all.csv', header=None)

        # 获取第二列并计算和（假设第二列的索引为1）
        total_power = df.iloc[:, 1].sum()
        res = np.append(res, total_power)

    print(f'均值：{res.mean()}')
    print(f'最大值：{res.max()}')
    print(f'最小值：{res.min()}')
